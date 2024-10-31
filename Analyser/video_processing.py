import os
import cv2
from collections import defaultdict
import openpyxl

def get_total_frames(video_path):
    cap = cv2.VideoCapture(video_path)
    prop = int(cv2.CAP_PROP_FRAME_COUNT)
    count = int(cv2.VideoCapture.get(cap, prop))
    return count

def process_video(model, video_path, output_folder, video_name, conf_threshold=0.6):
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print(f"Error: Unable to open video file {video_path}")
        return
    
    frame_idx = 1
    class_frames = defaultdict(list)
    frames_list = [''] * (get_total_frames(video_path) + 1)
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        results = model.predict(frame, conf=conf_threshold)
        detections = results[0].boxes

        for det in detections:
            class_id = int(det.cls)
            class_name = model.names[class_id]
            confidence = float(det.conf)
            class_frames[frame_idx].append((class_name, confidence))

        if not class_frames[frame_idx]:
            class_frames[frame_idx].append(("no_detection", 0))

        frame_idx += 1

    print("Capturing Done......")
    print("\n")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Behaviour Frames"
    
    header_row = ["Frame", "Behaviour", "Confidence"]
    for col, value in enumerate(header_row, start=1):
        ws.cell(row=1, column=col).value = value
        
    row_idx = 2
    for frame_idx, detections in class_frames.items():
        behaviour_str = ",".join([det[0] for det in detections])
        confidence_str = ",".join([f"{det[1]:.3f}" for det in detections])
        ws.cell(row=row_idx, column=1).value = frame_idx
        ws.cell(row=row_idx, column=2).value = behaviour_str
        ws.cell(row=row_idx, column=3).value = confidence_str
        row_idx += 1
        
    print("Total number of frames annotated are : ", row_idx)    

    ws.cell(row=row_idx, column=1).value = "Total number of frames annotated:"
    ws.cell(row=row_idx, column=2).value = row_idx 
    wb.save(f"{output_folder}/{video_name}_behaviour_frames.xlsx")
    print("Processing complete.")
